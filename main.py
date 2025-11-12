
from abc import ABC, abstractmethod
from datetime import datetime
import os
import io
import qrcode
from reportlab.lib.pagesizes import letter
import openpyxl



class InvoiceGenerator(ABC):
    def __init__(self, client_name, items, invoice_number=None):
        self.client_name = client_name
        self.items = items
        self.invoice_number = invoice_number or f"INV-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        self.date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def calculate_total(self):
        return sum(item['price'] for item in self.items)
    
    def generate_qr_data(self):
        return f"Invoice: {self.invoice_number}\nClient: {self.client_name}\nTotal: ${self.calculate_total():.2f}\nDate: {self.date}"
    
    @abstractmethod
    def generate_invoice(self):
        pass

class PDFInvoiceGenerator(InvoiceGenerator):
   
    
    def generate_invoice(self):
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            import qrcode
            
            filename = f"{self.invoice_number}.pdf"
            c = canvas.Canvas(filename, pagesize=letter)
            width, height = letter
            c.setFillColor(colors.HexColor('#2C3E50'))
            c.rect(0, height - 100, width, 100, fill=True, stroke=False)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 28)
            c.drawString(50, height - 60, "FinTrack Co.")
            
            c.setFont("Helvetica", 10)
            c.drawString(50, height - 80, "Professional Invoice Management")
            c.setFillColor(colors.HexColor('#2C3E50'))
            c.setFont("Helvetica-Bold", 24)
            c.drawString(50, height - 140, "INVOICE")
            
            c.setFont("Helvetica", 11)
            c.drawString(50, height - 165, f"Invoice #: {self.invoice_number}")
            c.drawString(50, height - 185, f"Date: {self.date}")
            
            c.setStrokeColor(colors.HexColor('#3498DB'))
            c.setLineWidth(2)
            c.rect(50, height - 260, 250, 50, fill=False, stroke=True)
            
            c.setFont("Helvetica-Bold", 12)
            c.drawString(60, height - 225, "BILL TO:")
            c.setFont("Helvetica", 11)
            c.drawString(60, height - 245, self.client_name)
        
            y = height - 310
            c.setFillColor(colors.HexColor('#3498DB'))
            c.rect(50, y - 25, width - 100, 25, fill=True, stroke=False)
            
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 12)
            c.drawString(60, y - 18, "Item Description")
            c.drawString(width - 150, y - 18, "Price")
            
        
            y -= 50
            c.setFillColor(colors.black)
            c.setFont("Helvetica", 11)
            
            for i, item in enumerate(self.items):
                if i % 2 == 0:
                    c.setFillColor(colors.HexColor('#ECF0F1'))
                    c.rect(50, y - 20, width - 100, 20, fill=True, stroke=False)
                
                c.setFillColor(colors.black)
                c.drawString(60, y - 12, item['name'])
                c.drawString(width - 150, y - 12, f"${item['price']:.2f}")
                y -= 25
            y -= 20
            c.setFillColor(colors.HexColor('#27AE60'))
            c.rect(width - 250, y - 35, 200, 35, fill=True, stroke=False)
            
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 16)
            c.drawString(width - 240, y - 22, "TOTAL:")
            c.drawString(width - 150, y - 22, f"${self.calculate_total():.2f}")
            qr = qrcode.QRCode(version=1, box_size=10, border=2)
            qr.add_data(self.generate_qr_data())
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
            qr_buffer = io.BytesIO()
            qr_img.save(qr_buffer, format='PNG')
            qr_buffer.seek(0)
            from reportlab.lib.utils import ImageReader
            qr_image = ImageReader(qr_buffer)
            c.drawImage(qr_image, width - 150, 100, width=100, height=100)
            
            c.setFillColor(colors.HexColor('#7F8C8D'))
            c.setFont("Helvetica", 8)
            c.drawString(width - 150, 80, "Scan for details")
            
            c.setFillColor(colors.HexColor('#95A5A6'))
            c.setFont("Helvetica", 9)
            c.drawString(50, 50, "Thank you for your business!")
            c.drawString(50, 35, "FinTrack Co. | www.fintrack.com | support@fintrack.com")
            
            c.save()
            return os.path.abspath(filename)
            
        except ImportError as e:
            print(f"⚠️ Kutubxona topilmadi: {e}")
            print("O'rnatish: pip install reportlab qrcode[pil]")
            return None



class ExcelInvoiceGenerator(InvoiceGenerator):
    
    def generate_invoice(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
            import qrcode
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice"
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            blue_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            green_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            gray_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
            
            white_font = Font(color="FFFFFF", bold=True, size=14)
            title_font = Font(bold=True, size=20, color="2C3E50")
            bold_font = Font(bold=True, size=12)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            ws.merge_cells('A1:E1')
            ws['A1'] = 'FinTrack Co.'
            ws['A1'].font = Font(bold=True, size=24, color="FFFFFF")
            ws['A1'].fill = header_fill
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40
            ws.merge_cells('A3:C3')
            ws['A3'] = 'INVOICE'
            ws['A3'].font = title_font
            
            ws['A5'] = 'Invoice #:'
            ws['B5'] = self.invoice_number
            ws['A6'] = 'Date:'
            ws['B6'] = self.date
            
            ws['A5'].font = bold_font
            ws['A6'].font = bold_font
            
          
            ws['A8'] = 'BILL TO:'
            ws['A8'].font = Font(bold=True, size=12, color="FFFFFF")
            ws['A8'].fill = blue_fill
            ws['A9'] = self.client_name
            ws['A9'].font = Font(size=11)
            headers = ['Item Description', 'Price']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=11, column=col)
                cell.value = header
                cell.font = white_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 15
            
          
            row = 12
            for i, item in enumerate(self.items):
                ws[f'A{row}'] = item['name']
                ws[f'B{row}'] = f"${item['price']:.2f}"
                
                
                if i % 2 == 0:
                    ws[f'A{row}'].fill = gray_fill
                    ws[f'B{row}'].fill = gray_fill
                
                ws[f'A{row}'].border = thin_border
                ws[f'B{row}'].border = thin_border
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                
                row += 1
            
         
            row += 1
            ws[f'A{row}'] = 'TOTAL'
            ws[f'B{row}'] = f"${self.calculate_total():.2f}"
            ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f'B{row}'].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f'A{row}'].fill = green_fill
            ws[f'B{row}'].fill = green_fill
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            qr = qrcode.QRCode(version=1, box_size=10, border=2)
            qr.add_data(self.generate_qr_data())
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
            qr_path = f"temp_qr_{self.invoice_number}.png"
            qr_img.save(qr_path)
            
            img = XLImage(qr_path)
            img.width = 150
            img.height = 150
            ws.add_image(img, f'D5')
            
            # 📌 FOOTER
            row += 3
            ws[f'A{row}'] = 'Thank you for your business!'
            ws[f'A{row}'].font = Font(italic=True, color="7F8C8D")
            
            filename = f"{self.invoice_number}.xlsx"
            wb.save(filename)
            if os.path.exists(qr_path):
                os.remove(qr_path)
            
            return os.path.abspath(filename)
            
        except ImportError as e:
            print(f"⚠️ Kutubxona topilmadi: {e}")
            print("O'rnatish: pip install openpyxl qrcode[pil]")
            return None

class HTMLInvoiceGenerator(InvoiceGenerator):
    
    def generate_invoice(self):
        try:
            import qrcode
            import base64
          
            qr = qrcode.QRCode(version=1, box_size=10, border=2)
            qr.add_data(self.generate_qr_data())
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")
            
          
            buffered = io.BytesIO()
            qr_img.save(buffered, format="PNG")
            qr_base64 = base64.b64encode(buffered.getvalue()).decode()
            
           
            items_html = ""
            for i, item in enumerate(self.items):
                bg_class = "bg-gray" if i % 2 == 0 else ""
                items_html += f"""
                <tr class="{bg_class}">
                    <td>{item['name']}</td>
                    <td>${item['price']:.2f}</td>
                </tr>
                """
            
            html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Invoice - {self.invoice_number}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 40px 20px;
            min-height: 100vh;
        }}
        
        .container {{
            max-width: 900px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #2C3E50 0%, #34495e 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 36px;
            margin-bottom: 10px;
            letter-spacing: 2px;
        }}
        
        .header p {{
            font-size: 14px;
            opacity: 0.9;
        }}
        
        .content {{
            padding: 40px;
        }}
        
        .invoice-info {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 30px;
            margin-bottom: 40px;
        }}
        
        .info-box {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 10px;
            border-left: 4px solid #3498DB;
        }}
        
        .info-box h3 {{
            color: #2C3E50;
            margin-bottom: 10px;
            font-size: 18px;
        }}
        
        .info-box p {{
            color: #7f8c8d;
            font-size: 14px;
            line-height: 1.6;
        }}
        
        .client-box {{
            border-left-color: #9b59b6;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 30px 0;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            border-radius: 10px;
            overflow: hidden;
        }}
        
        thead {{
            background: linear-gradient(135deg, #3498DB 0%, #2980b9 100%);
            color: white;
        }}
        
        th {{
            padding: 20px;
            text-align: left;
            font-weight: 600;
            font-size: 14px;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}
        
        td {{
            padding: 18px 20px;
            border-bottom: 1px solid #ecf0f1;
            font-size: 15px;
        }}
        
        tr.bg-gray {{
            background-color: #f8f9fa;
        }}
        
        tr:hover {{
            background-color: #e8f4f8;
            transition: all 0.3s;
        }}
        
        .total-row {{
            background: linear-gradient(135deg, #27AE60 0%, #229954 100%);
            color: white;
            font-weight: bold;
            font-size: 18px;
        }}
        
        .total-row td {{
            padding: 25px 20px;
            border: none;
        }}
        
        .qr-section {{
            text-align: center;
            margin: 40px 0;
            padding: 30px;
            background: #f8f9fa;
            border-radius: 10px;
        }}
        
        .qr-section img {{
            border: 3px solid #3498DB;
            border-radius: 10px;
            padding: 10px;
            background: white;
        }}
        
        .qr-section p {{
            margin-top: 15px;
            color: #7f8c8d;
            font-size: 14px;
        }}
        
        .footer {{
            background: #2C3E50;
            color: white;
            padding: 30px 40px;
            text-align: center;
        }}
        
        .footer p {{
            margin: 5px 0;
            font-size: 14px;
        }}
        
        .footer .thank-you {{
            font-size: 18px;
            font-weight: 600;
            margin-bottom: 15px;
        }}
        
        @media print {{
            body {{ background: white; padding: 0; }}
            .container {{ box-shadow: none; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <!-- HEADER -->
        <div class="header">
            <h1>🏢 FinTrack Co.</h1>
            <p>Professional Invoice Management System</p>
        </div>
        
        <!-- CONTENT -->
        <div class="content">
            <!-- INVOICE INFO -->
            <div class="invoice-info">
                <div class="info-box">
                    <h3>📄 Invoice Details</h3>
                    <p><strong>Invoice #:</strong> {self.invoice_number}</p>
                    <p><strong>Date:</strong> {self.date}</p>
                </div>
                
                <div class="info-box client-box">
                    <h3>👤 Bill To</h3>
                    <p><strong>{self.client_name}</strong></p>
                </div>
            </div>
            
            <!-- TABLE -->
            <table>
                <thead>
                    <tr>
                        <th>Item Description</th>
                        <th style="text-align: right;">Price</th>
                    </tr>
                </thead>
                <tbody>
                    {items_html}
                    <tr class="total-row">
                        <td><strong>TOTAL</strong></td>
                        <td style="text-align: right;"><strong>${self.calculate_total():.2f}</strong></td>
                    </tr>
                </tbody>
            </table>
            
            <!-- QR CODE -->
            <div class="qr-section">
                <img src="data:image/png;base64,{qr_base64}" alt="QR Code" width="200" height="200">
                <p>📱 Scan this QR code for invoice verification</p>
            </div>
        </div>
        
        <!-- FOOTER -->
        <div class="footer">
            <p class="thank-you">✨ Thank you for your business!</p>
            <p>FinTrack Co. | www.fintrack.com | support@fintrack.com</p>
            <p>© 2025 FinTrack Co. All rights reserved.</p>
        </div>
    </div>
</body>
</html>
            """
            
            filename = f"{self.invoice_number}.html"
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return os.path.abspath(filename)
            
        except ImportError as e:
            print(f"⚠️ Kutubxona topilmadi: {e}")
            print("O'rnatish: pip install qrcode[pil]")
            return None



class InvoiceManager:
    
    def __init__(self, generator):
        self.generator = generator
    
    def export_invoice(self):
        """Invoice yaratish va fayl yo'lini ko'rsatish"""
        format_name = self.generator.__class__.__name__.replace('Generator', '')
        print(f"\n📄 {format_name} yaratilmoqda...")
        
        filepath = self.generator.generate_invoice()
        
        if filepath:
            print(f"✅ Tayyor: {filepath}")
            return filepath
        else:
            print("❌ Xatolik yuz berdi")
            return None



if __name__ == "__main__":
    print("=" * 70)
    print("🏢 FinTrack Co. - Professional Invoice Generator")
    print("=" * 70)
    
  
    if not os.path.exists('invoices'):
        os.makedirs('invoices')
        print("📁 'invoices' papkasi yaratildi\n")
    
    os.chdir('invoices')
    
  
    client = "Acme Corporation"
    items = [
        {'name': 'Website Development & Design', 'price': 2500.00},
        {'name': 'Mobile App Development', 'price': 3500.00},
        {'name': 'SEO & Digital Marketing', 'price': 1200.00},
        {'name': 'Cloud Hosting (12 months)', 'price': 600.00},
        {'name': 'Technical Support Package', 'price': 800.00}
    ]
    
    print("🎯 Invoice yaratish boshlandi...\n")
    
    invoice_num = f"INV-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    
    # 1. PDF Invoice
    pdf_gen = PDFInvoiceGenerator(client, items, invoice_num)
    manager = InvoiceManager(pdf_gen)
    manager.export_invoice()
    
    # 2. Excel Invoice  
    excel_gen = ExcelInvoiceGenerator(client, items, invoice_num)
    manager = InvoiceManager(excel_gen)
    manager.export_invoice()
    
    # 3. HTML Invoice
    html_gen = HTMLInvoiceGenerator(client, items, invoice_num)
    manager = InvoiceManager(html_gen)
    manager.export_invoice()
    
    print("\n" + "=" * 70)
    print("✨ Barcha invoicelar muvaffaqiyatli yaratildi!")
    print(f"📁 Fayl manzili: {os.path.abspath('.')}")
    print("=" * 70)